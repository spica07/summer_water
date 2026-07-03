# -*- coding: utf-8 -*-
"""verify_result_*.json 의 수정사항을 xlsx(전체목록 시트)에 반영"""
import html
import json
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

TOOLS = Path(__file__).resolve().parent
BASE = TOOLS.parent
XLSX = BASE / "seoul_public_water_playgrounds_2026.xlsx"
TODAY = "2026-07-03"

FIELD_COL = {
    "statusRaw": "2026 운영상태",
    "period": "운영기간",
    "hours": "운영시간",
    "closedInfo": "휴장/중단",
    "feeInfo": "대상/요금",
    "reservation": "예약",
    "address": "주소/위치",
    "source": "출처명",
    "sourceUrl": "출처URL",
}

results = []
for i in range(1, 5):
    p = TOOLS / f"verify_result_{i}.json"
    results.append(json.loads(p.read_text(encoding="utf-8")))

wb = openpyxl.load_workbook(XLSX)
ws = wb["전체목록"]
header = {str(c.value).strip(): idx + 1 for idx, c in enumerate(ws[1])}

def row_of(fid):
    return fid + 1  # id는 데이터행 순번(1..114), 헤더가 1행

applied, skipped = [], []
touched_rows = set()

for res in results:
    for c in res["corrections"]:
        fid, field = c["id"], c["field"]
        if field not in FIELD_COL:
            skipped.append((fid, field, "unknown field"))
            continue
        r = row_of(fid)
        sheet_name = str(ws.cell(row=r, column=header["시설명"]).value or "").strip()
        # 시설명 대조(부분일치 허용: 에이전트가 괄호 병기 등을 붙이는 경우)
        agent_name = c["name"].split("(")[0].strip()
        if agent_name[:6] not in sheet_name and sheet_name[:6] not in c["name"]:
            skipped.append((fid, field, f"name mismatch: sheet={sheet_name!r} agent={c['name']!r}"))
            continue
        new_val = html.unescape(c["corrected"]) if isinstance(c["corrected"], str) else c["corrected"]
        cell = ws.cell(row=r, column=header[FIELD_COL[field]])
        old_val = cell.value
        if (old_val or None) == (new_val or None):
            continue
        cell.value = new_val
        touched_rows.add(r)
        applied.append((fid, sheet_name, field, old_val, new_val))

    # 2026 운영이 공식 확인된 시설은 검증등급 A로 상향
    for fid in res.get("confirmed2026", []):
        r = row_of(fid)
        cell = ws.cell(row=r, column=header["검증등급"])
        if cell.value != "A":
            applied.append((fid, ws.cell(row=r, column=header["시설명"]).value, "검증등급", cell.value, "A"))
            cell.value = "A"
            touched_rows.add(r)

# 수정된 행의 조사일 갱신
for r in touched_rows:
    ws.cell(row=r, column=header["조사일"]).value = TODAY

# 자치구별_요약: 최고 검증등급 재계산
ws2 = wb["자치구별_요약"]
h2 = {str(c.value).strip(): idx + 1 for idx, c in enumerate(ws2[1])}
if "최고 검증등급" in h2 and "자치구" in h2:
    best = {}
    for r in range(2, ws.max_row + 1):
        d = ws.cell(row=r, column=header["자치구"]).value
        g = ws.cell(row=r, column=header["검증등급"]).value
        if d and g:
            best[d] = min(best.get(d, "Z"), str(g))
    for r in range(2, ws2.max_row + 1):
        d = ws2.cell(row=r, column=h2["자치구"]).value
        if d in best:
            ws2.cell(row=r, column=h2["최고 검증등급"]).value = best[d]

wb.save(XLSX)

print(f"applied: {len(applied)} changes across {len(touched_rows)} rows")
for a in applied:
    print(f"  [id {a[0]:3d}] {a[1]} | {a[2]}: {str(a[3])[:40]!r} -> {str(a[4])[:60]!r}")
if skipped:
    print(f"\nskipped: {len(skipped)}")
    for s in skipped:
        print(" ", s)
