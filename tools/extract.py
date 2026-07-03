# -*- coding: utf-8 -*-
"""xlsx -> facilities_raw.json 추출 + 필드 정규화"""
import json
import re
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

BASE = Path(__file__).resolve().parent.parent
XLSX = BASE / "seoul_public_water_playgrounds_2026.xlsx"
OUT = Path(__file__).resolve().parent / "facilities_raw.json"

wb = openpyxl.load_workbook(XLSX, data_only=True)

# ---------- 전체목록 ----------
ws = wb["전체목록"]
rows = list(ws.iter_rows(values_only=True))
header = [str(h).strip() if h is not None else "" for h in rows[0]]
print("columns:", header)

def cell(row, col_name):
    idx = header.index(col_name)
    v = row[idx]
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None

import calendar

FULL_RANGE_RE = re.compile(
    r"(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})\s*[~\-～]\s*(?:(\d{4})[.\-/])?(\d{1,2})[.\-/](\d{1,2})"
)
MONTH_RANGE_RE = re.compile(r"(\d{4})[.\-/](\d{1,2})월?\s*[~\-～]\s*(\d{1,2})월?(?![.\-/\d])")
END_ONLY_RE = re.compile(r"^\s*[~\-～]\s*(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})|(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})\s*까지")
START_ONLY_RE = re.compile(r"(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})\s*(?:개장|부터|[~\-～]\s*$|[~\-～]\s*[^\d])")
SINGLE_DATE_RE = re.compile(r"(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})")
MONTH_ONLY_RE = re.compile(r"(\d{4})[.\-/](\d{1,2})월?\s*운영")

def _fmt(y, m, d):
    return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"

def _month_end(y, m):
    return _fmt(y, m, calendar.monthrange(int(y), int(m))[1])

def parse_period(s):
    """운영기간 문자열 -> (start, end). 2026년 정보만 파싱하고 과거 연도는 참고치로 남긴다."""
    if not s:
        return None, None
    start = end = None
    m = FULL_RANGE_RE.search(s)
    if m:
        y1, mo1, d1, y2, mo2, d2 = m.groups()
        start, end = _fmt(y1, mo1, d1), _fmt(y2 or y1, mo2, d2)
    else:
        m = MONTH_RANGE_RE.search(s)
        if m:
            y, mo1, mo2 = m.groups()
            start, end = _fmt(y, mo1, 1), _month_end(y, mo2)
        else:
            m = END_ONLY_RE.search(s)
            if m:
                g = m.groups()
                y, mo, d = (g[0], g[1], g[2]) if g[0] else (g[3], g[4], g[5])
                end = _fmt(y, mo, d)
            else:
                m = START_ONLY_RE.search(s) or (SINGLE_DATE_RE.search(s) if "개장" in s or s.strip().endswith("~") else None)
                if m:
                    y, mo, d = m.group(1), m.group(2), m.group(3)
                    start = _fmt(y, mo, d)
                else:
                    m = MONTH_ONLY_RE.search(s)
                    if m:
                        y, mo = m.groups()
                        start, end = _fmt(y, mo, 1), _month_end(y, mo)
    # 2025년 이하 정보는 전년도 참고치 -> 날짜 필터에 쓰지 않음
    ref_year = start or end
    if ref_year and int(ref_year[:4]) < 2026:
        return None, None
    return start, end

def norm_status(status_raw, closed_raw):
    s = (status_raw or "")
    if re.search(r"중단|취소|폐쇄", s):
        return "중단"
    if re.search(r"운영\s*중", s):
        return "운영중"
    if re.search(r"예정", s):
        return "운영예정"
    if re.search(r"미공지|미발표", s):
        return "전년도 확인" if re.search(r"2025|전년도", s) else "재확인 필요"
    if re.search(r"2026", s):
        return "2026 확인"
    if re.search(r"2025|전년도", s):
        return "전년도 확인"
    if re.search(r"재확인|미정|검토", s):
        return "재확인 필요"
    if s:
        return "기타 확인"
    return "재확인 필요"

def norm_type(t):
    if not t:
        return "기타"
    if re.search(r"한강", t):
        return "한강 물놀이장"
    if re.search(r"수영장", t):
        return "야외 수영장"
    if re.search(r"분수", t):
        return "물놀이 분수"
    if re.search(r"계곡|하천|천변", t):
        return "하천/계곡"
    if re.search(r"물놀이", t):
        return "공원 물놀이터"
    return "기타"

def parse_free(fee_raw):
    if not fee_raw:
        return None
    has_free = re.search(r"무료", fee_raw) is not None
    has_paid = re.search(r"유료|\d[\d,]*\s*원", fee_raw) is not None
    if has_free and not has_paid:
        return True
    if has_paid:
        return False
    return None

def parse_reservation(r):
    if not r:
        return None
    if re.search(r"필수|사전\s*예약|예약제", r):
        return True
    if re.search(r"불필요|없음|현장|자유|해당\s*없", r):
        return False
    return None

facilities = []
for i, row in enumerate(rows[1:], start=1):
    if all(v is None for v in row):
        continue
    status_raw = cell(row, "2026 운영상태")
    closed_raw = cell(row, "휴장/중단")
    period_raw = cell(row, "운영기간")
    fee_raw = cell(row, "대상/요금")
    resv_raw = cell(row, "예약")
    type_raw = cell(row, "시설유형")
    ps, pe = parse_period(period_raw)
    facilities.append({
        "id": i,
        "district": cell(row, "자치구"),
        "name": cell(row, "시설명"),
        "type": norm_type(type_raw),
        "typeRaw": type_raw,
        "address": cell(row, "주소/위치"),
        "status": norm_status(status_raw, closed_raw),
        "statusRaw": status_raw,
        "period": period_raw,
        "periodStart": ps,
        "periodEnd": pe,
        "hours": cell(row, "운영시간"),
        "closedInfo": closed_raw,
        "feeInfo": fee_raw,
        "isFree": parse_free(fee_raw),
        "reservation": resv_raw,
        "needsReservation": parse_reservation(resv_raw),
        "grade": cell(row, "검증등급"),
        "source": cell(row, "출처명"),
        "sourceUrl": cell(row, "출처URL"),
        "note": cell(row, "비고"),
    })

# ---------- 자치구별_요약 ----------
ws2 = wb["자치구별_요약"]
rows2 = list(ws2.iter_rows(values_only=True))
h2 = [str(h).strip() if h is not None else "" for h in rows2[0]]
summary = []
for row in rows2[1:]:
    if row[0] is None:
        continue
    summary.append({h2[j]: (str(v).strip() if v is not None else None) for j, v in enumerate(row)})

OUT.write_text(
    json.dumps({"facilities": facilities, "summary": summary}, ensure_ascii=False, indent=2),
    encoding="utf-8",
)

# ---------- 리포트 ----------
from collections import Counter
print(f"\ntotal facilities: {len(facilities)}")
for field in ["status", "type", "grade", "district"]:
    print(f"\n[{field}]")
    for k, v in Counter(f[field] for f in facilities).most_common():
        print(f"  {k}: {v}")
print("\n[isFree]", Counter(str(f["isFree"]) for f in facilities))
print("[needsReservation]", Counter(str(f["needsReservation"]) for f in facilities))
print("[periodStart parsed]", sum(1 for f in facilities if f["periodStart"]), "/", len(facilities))
print("[address missing]", sum(1 for f in facilities if not f["address"]))
print("\n-- distinct statusRaw --")
for k, v in Counter((f["statusRaw"] or "") for f in facilities).most_common():
    print(f"  {k!r}: {v}")
print("\n-- distinct typeRaw --")
for k, v in Counter((f["typeRaw"] or "") for f in facilities).most_common():
    print(f"  {k!r}: {v}")
